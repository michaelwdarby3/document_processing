from __future__ import annotations

from dataclasses import dataclass
import json
from pathlib import Path
from typing import List

import openpyxl

from docling_offline.utils import (
    clip_overlapping_table_cells,
    write_table_metadata,
    export_tables_to_xlsx,
    sanitize_table_bboxes,
    detect_garbled_tables,
    merge_repaired_tables,
    _table_title,
)
from docling_core.types.doc.base import CoordOrigin


@dataclass
class StubBBox:
    l: float
    t: float
    r: float
    b: float
    coord_origin: CoordOrigin = CoordOrigin.TOPLEFT


@dataclass
class StubProv:
    bbox: StubBBox
    page_no: int


@dataclass
class StubCell:
    bbox: StubBBox
    start_col_offset_idx: int
    end_col_offset_idx: int
    text: str = ""


@dataclass
class StubText:
    bbox: StubBBox
    page_no: int
    text: str
    label: object = "section_header"

    def __post_init__(self) -> None:
        self.prov = [StubProv(bbox=self.bbox, page_no=self.page_no)]


@dataclass
class StubTableData:
    table_cells: List[StubCell]
    num_cols: int

    @property
    def grid(self):
        rows: List[List[StubCell]] = []
        row: List[StubCell] = []
        for idx, cell in enumerate(self.table_cells):
            if idx % max(self.num_cols, 1) == 0 and row:
                rows.append(row)
                row = []
            row.append(cell)
        if row:
            rows.append(row)
        return rows


class StubTable:
    def __init__(self, bbox: StubBBox, page_no: int, cells: List[StubCell], num_cols: int) -> None:
        self.prov = [StubProv(bbox=bbox, page_no=page_no)]
        self.data = StubTableData(table_cells=cells, num_cols=num_cols)


class StubSize:
    def __init__(self, width: float, height: float) -> None:
        self.width = width
        self.height = height


class StubPage:
    def __init__(self, width: float = 600.0, height: float = 800.0) -> None:
        self.size = StubSize(width, height)


class StubDoc:
    def __init__(self, tables, page_no: int = 1, texts: List[StubText] | None = None) -> None:
        self.tables = tables
        self.pages = {page_no: StubPage()}
        self.texts = texts or []


class FakeLabel:
    def __init__(self, value: str) -> None:
        self.value = value


def test_clip_overlapping_table_cells_removes_left_bleed() -> None:
    left_table = StubTable(
        bbox=StubBBox(l=10, t=500, r=100, b=100),
        page_no=1,
        cells=[],
        num_cols=0,
    )
    overlapping_table = StubTable(
        bbox=StubBBox(l=80, t=400, r=220, b=120),
        page_no=1,
        cells=[
            StubCell(bbox=StubBBox(l=60, t=390, r=90, b=360), start_col_offset_idx=0, end_col_offset_idx=1, text="bleed"),
            StubCell(bbox=StubBBox(l=140, t=390, r=210, b=360), start_col_offset_idx=1, end_col_offset_idx=2, text="main"),
        ],
        num_cols=2,
    )
    doc = StubDoc([left_table, overlapping_table])

    removed, per_table = clip_overlapping_table_cells(doc, gutter=0.0)

    assert removed == 1
    assert per_table
    assert overlapping_table.data.num_cols == 1
    assert len(overlapping_table.data.table_cells) == 1
    assert overlapping_table.data.table_cells[0].text == "main"
    assert overlapping_table.data.table_cells[0].start_col_offset_idx == 0


def test_clip_overlapping_table_cells_is_noop_without_overlap() -> None:
    first = StubTable(
        bbox=StubBBox(l=10, t=500, r=100, b=100),
        page_no=1,
        cells=[],
        num_cols=0,
    )
    second = StubTable(
        bbox=StubBBox(l=200, t=400, r=320, b=120),
        page_no=1,
        cells=[
            StubCell(bbox=StubBBox(l=210, t=390, r=260, b=360), start_col_offset_idx=0, end_col_offset_idx=1, text="kept")
        ],
        num_cols=1,
    )
    doc = StubDoc([first, second])

    removed, per_table = clip_overlapping_table_cells(doc)

    assert removed == 0
    assert per_table == {}
    assert second.data.num_cols == 1
    assert len(second.data.table_cells) == 1


def test_write_table_metadata(tmp_path: Path) -> None:
    table = StubTable(
        bbox=StubBBox(l=10, t=50, r=60, b=0, coord_origin=CoordOrigin.BOTTOMLEFT),
        page_no=1,
        cells=[
            StubCell(
                bbox=StubBBox(l=10, t=40, r=20, b=30, coord_origin=CoordOrigin.BOTTOMLEFT),
                start_col_offset_idx=0,
                end_col_offset_idx=1,
                text="A",
            )
        ],
        num_cols=1,
    )
    heading = StubText(
        bbox=StubBBox(l=5, t=90, r=65, b=70, coord_origin=CoordOrigin.BOTTOMLEFT),
        page_no=1,
        text="Demo Heading",
    )
    doc = StubDoc([table], texts=[heading])
    target = tmp_path / "demo.tables.json"

    write_table_metadata(
        doc,
        target,
        clipping_enabled=True,
        clipped_cells_total=2,
        clipped_by_table={id(table): 2},
        table_mode="accurate",
        table_cell_matching=False,
        formats=("json", "xlsx"),
        page_metadata=sanitize_table_bboxes(doc),
        garbled_tables={0},
        repaired_tables={0},
    )

    data = json.loads(target.read_text())
    assert data["document"]["page_count"] == 1
    assert data["table_processing"]["clip_table_overlap"] is True
    assert data["table_processing"]["clipped_cells_total"] == 2
    assert data["table_processing"]["tables_flagged_garbled"] == 1
    assert data["table_processing"]["tables_repaired_with_ocr"] == 1
    assert data["tables"][0]["metrics"]["num_cols"] == 1
    assert data["tables"][0]["clipped_cells"] == 2
    assert data["tables"][0]["flags"]["garbled_input"] is True
    assert data["tables"][0]["flags"]["repaired_with_ocr"] is True
    assert data["tables"][0]["title"] == "Demo Heading"


def test_export_tables_to_xlsx(tmp_path: Path) -> None:
    table = StubTable(
        bbox=StubBBox(l=0, t=10, r=20, b=0, coord_origin=CoordOrigin.BOTTOMLEFT),
        page_no=1,
        cells=[
            StubCell(
                bbox=StubBBox(l=0, t=10, r=10, b=0, coord_origin=CoordOrigin.BOTTOMLEFT),
                start_col_offset_idx=0,
                end_col_offset_idx=1,
                text="A",
            ),
            StubCell(
                bbox=StubBBox(l=10, t=10, r=20, b=0, coord_origin=CoordOrigin.BOTTOMLEFT),
                start_col_offset_idx=1,
                end_col_offset_idx=2,
                text="=SUM(1,2)",
            ),
        ],
        num_cols=2,
    )
    heading = StubText(
        bbox=StubBBox(l=0, t=20, r=20, b=12, coord_origin=CoordOrigin.BOTTOMLEFT),
        page_no=1,
        text="Components Title",
    )
    doc = StubDoc([table], texts=[heading])
    target = tmp_path / "demo.xlsx"

    created = export_tables_to_xlsx(doc, target)

    assert created == target
    assert target.exists()
    wb = openpyxl.load_workbook(target, read_only=True)
    sheet_name = wb.sheetnames[0]
    assert sheet_name == "P1_T01_Components Title"
    sheet = wb[sheet_name]
    first_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert first_row[0] == "A"
    assert first_row[1].startswith("\u200b") or first_row[1].startswith("\u200B")
    assert first_row[1].strip("\u200b\u200B") == "=SUM(1,2)"
    wb.close()


def test_table_title_skips_garbled_heading() -> None:
    table = StubTable(
        bbox=StubBBox(l=0, t=10, r=60, b=0, coord_origin=CoordOrigin.BOTTOMLEFT),
        page_no=1,
        cells=[],
        num_cols=1,
    )
    heading = StubText(
        bbox=StubBBox(l=0, t=40, r=10, b=20, coord_origin=CoordOrigin.BOTTOMLEFT),
        page_no=1,
        text="\u000b\u0000??",
    )
    doc = StubDoc([table], texts=[heading])

    assert _table_title(table, 1, doc) == "Table 01"


def test_table_title_uses_enum_heading_label() -> None:
    table = StubTable(
        bbox=StubBBox(l=0, t=10, r=60, b=0, coord_origin=CoordOrigin.BOTTOMLEFT),
        page_no=2,
        cells=[],
        num_cols=1,
    )
    heading = StubText(
        bbox=StubBBox(l=0, t=80, r=60, b=60, coord_origin=CoordOrigin.BOTTOMLEFT),
        page_no=2,
        text="Enum Heading Title",
        label=FakeLabel("section_header"),
    )
    doc = StubDoc([table], page_no=2, texts=[heading])

    assert _table_title(table, 1, doc) == "Enum Heading Title"


def test_sanitize_table_bboxes_clamps_coordinates() -> None:
    table = StubTable(
        bbox=StubBBox(l=0, t=820, r=100, b=780),
        page_no=1,
        cells=[
            StubCell(
                bbox=StubBBox(l=-10, t=820, r=610, b=-5),
                start_col_offset_idx=0,
                end_col_offset_idx=1,
            )
        ],
        num_cols=1,
    )
    doc = StubDoc([table])

    meta = sanitize_table_bboxes(doc)

    cell_bbox = table.data.table_cells[0].bbox
    assert cell_bbox.l == 0
    assert cell_bbox.r == 600
    assert 0 <= cell_bbox.t <= 800
    assert 0 <= cell_bbox.b <= 800
    assert meta[1]["width"] == 600


def test_detect_garbled_tables_flags_symbols() -> None:
    noisy_text = "!!@@###$$$$%%%^^^^&&&***(((" * 20
    noisy_cell = StubCell(bbox=StubBBox(l=0, t=10, r=10, b=0), start_col_offset_idx=0, end_col_offset_idx=1, text=noisy_text)
    clean_cell = StubCell(bbox=StubBBox(l=0, t=10, r=10, b=0), start_col_offset_idx=0, end_col_offset_idx=1, text="Legible text value 123")
    noisy_table = StubTable(bbox=StubBBox(l=0, t=10, r=10, b=0), page_no=1, cells=[noisy_cell], num_cols=1)
    clean_table = StubTable(bbox=StubBBox(l=0, t=10, r=10, b=0), page_no=1, cells=[clean_cell], num_cols=1)
    doc = StubDoc([noisy_table, clean_table])

    suspect_map, suspect_indexes = detect_garbled_tables(doc)

    assert 0 in suspect_indexes
    assert 1 not in suspect_indexes
    assert 1 in suspect_map  # page key is 1-based


def test_merge_repaired_tables_prefers_high_iou() -> None:
    original_cell = StubCell(
        bbox=StubBBox(l=0, t=10, r=20, b=0),
        start_col_offset_idx=0,
        end_col_offset_idx=1,
        text="garbled",
    )
    repaired_cell = StubCell(
        bbox=StubBBox(l=0, t=10, r=20, b=0),
        start_col_offset_idx=0,
        end_col_offset_idx=1,
        text="fixed text",
    )
    original = StubTable(bbox=StubBBox(l=0, t=10, r=20, b=0), page_no=1, cells=[original_cell], num_cols=1)
    repaired = StubTable(bbox=StubBBox(l=1, t=10, r=21, b=0), page_no=1, cells=[repaired_cell], num_cols=1)
    base_doc = StubDoc([original])
    repair_doc = StubDoc([repaired])

    replaced = merge_repaired_tables(base_doc, repair_doc, [0], 1)

    assert replaced == {0}
    assert base_doc.tables[0].data.table_cells[0].text == "fixed text"
