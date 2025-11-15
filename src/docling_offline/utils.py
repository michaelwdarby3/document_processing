from __future__ import annotations

import itertools
import json
import logging
import math
import re
import shutil
import string
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

try:  # pragma: no cover - docling_core is a runtime dependency
    from docling_core.types.doc import TableItem
    from docling_core.types.doc.base import BoundingBox, CoordOrigin
except ImportError:  # pragma: no cover
    TableItem = Any  # type: ignore
    BoundingBox = Any  # type: ignore
    CoordOrigin = Any  # type: ignore

logger = logging.getLogger(__name__)


def ensure_tableformer_structure(artifacts_path: Path) -> None:
    tableformer_root = artifacts_path / "model_artifacts" / "tableformer"
    if not tableformer_root.exists():
        return

    _ensure_mode(tableformer_root, "fast", ["fast", "fat"])
    _ensure_mode(tableformer_root, "accurate", ["accurate", "acr"])


def clip_overlapping_table_cells(document: Any, gutter: float = 1.0) -> Tuple[int, Dict[int, int]]:
    """Remove stray TableFormer cells that spill into neighbouring tables."""

    tables: Sequence[TableItem] = getattr(document, "tables", None) or []
    pages: Any = getattr(document, "pages", None)
    if not tables or pages is None:
        return 0, {}

    tables_by_page: dict[int, List[Tuple[TableItem, BoundingBox]]] = defaultdict(list)
    for table in tables:
        bbox = _table_bbox(table)
        page_no = _table_page(table)
        if bbox is None or page_no is None:
            continue
        tables_by_page[page_no].append((table, bbox))

    removed = 0
    removed_by_table: Dict[int, int] = {}
    for page_tables in tables_by_page.values():
        page_tables.sort(key=lambda item: _normalize_bounds(item[1])[0])
        for idx, (table, bbox) in enumerate(page_tables):
            bounds = _normalize_bounds(bbox)
            clip_left = bounds[0]
            for prev_idx in range(idx):
                prev_table, prev_bbox = page_tables[prev_idx]
                prev_bounds = _normalize_bounds(prev_bbox)
                if (
                    _vertical_overlap(prev_bounds, bounds)
                    and prev_bounds[1] > clip_left
                ):
                    clip_left = max(clip_left, prev_bounds[1] + gutter)
            if clip_left > bounds[0]:
                trimmed = _clip_cells_left_of(table, clip_left)
                if trimmed:
                    removed += trimmed
                    removed_by_table[id(table)] = removed_by_table.get(id(table), 0) + trimmed
    return removed, removed_by_table


def sanitize_table_bboxes(document: Any) -> Dict[int, dict[str, float]]:
    """Clamp all table bounding boxes to their page dimensions."""

    page_meta = _page_metadata(document)
    tables: Sequence[TableItem] = getattr(document, "tables", None) or []
    for table in tables:
        page_no = _table_page(table)
        stats = page_meta.get(page_no) or {}
        page_w = stats.get("width")
        page_h = stats.get("height")
        if page_w is None or page_h is None:
            continue
        data = getattr(table, "data", None)
        cells: Sequence[Any] = getattr(data, "table_cells", None) or []
        for cell in cells:
            bbox = getattr(cell, "bbox", None)
            if bbox is None:
                continue
            _clamp_bbox(bbox, page_w, page_h)
    return page_meta


PRINTABLE_THRESHOLD = 0.7
ALNUM_THRESHOLD = 0.35
LOW_ALNUM_THRESHOLD = 0.05
PUNCT_THRESHOLD = 0.4
ALPHA_THRESHOLD = 0.15
MIN_TEXT_LENGTH = 40
TITLE_PRINTABLE_THRESHOLD = 0.6
TITLE_ALNUM_THRESHOLD = 0.15
TITLE_ALPHA_THRESHOLD = 0.1
TITLE_MIN_ALPHA = 3
HEADING_LABELS = {"section_header", "title", "heading", "text"}
_CONTROL_CHARS = set(chr(i) for i in range(0, 32)) - {"\n", "\r", "\t"}


def detect_garbled_tables(document: Any) -> Tuple[Dict[int, List[int]], Set[int]]:
    """Identify tables whose text content looks like corrupted encodings."""

    suspect_by_page: Dict[int, List[int]] = defaultdict(list)
    suspect_indexes: Set[int] = set()
    tables: Sequence[TableItem] = getattr(document, "tables", None) or []
    for idx, table in enumerate(tables):
        sample = _table_text_sample(table)
        cleaned = sample.replace("\n", " ").replace("\r", " ").strip()
        if len(cleaned) < MIN_TEXT_LENGTH:
            continue
        printable = sum(1 for ch in cleaned if ch.isprintable())
        alnum = sum(1 for ch in cleaned if ch.isalnum())
        control = sum(1 for ch in cleaned if ch in _CONTROL_CHARS)
        alpha = sum(1 for ch in cleaned if ch.isalpha())
        punct = sum(1 for ch in cleaned if ch in string.punctuation)
        ratio_printable = printable / len(cleaned)
        ratio_alnum = alnum / len(cleaned)
        control_ratio = control / len(cleaned)
        alpha_ratio = alpha / len(cleaned)
        punct_ratio = punct / len(cleaned)
        if ratio_alnum < LOW_ALNUM_THRESHOLD:
            page_no = _table_page(table)
            if page_no is None:
                continue
            suspect_by_page[page_no].append(idx)
            suspect_indexes.add(idx)
            continue
        if ratio_alnum < ALNUM_THRESHOLD:
            page_no = _table_page(table)
            if page_no is None:
                continue
            suspect_by_page[page_no].append(idx)
            suspect_indexes.add(idx)
            continue
        if alpha_ratio < ALPHA_THRESHOLD and punct_ratio > PUNCT_THRESHOLD:
            page_no = _table_page(table)
            if page_no is None:
                continue
            suspect_by_page[page_no].append(idx)
            suspect_indexes.add(idx)
            continue
        if ratio_printable < PRINTABLE_THRESHOLD:
            page_no = _table_page(table)
            if page_no is None:
                continue
            suspect_by_page[page_no].append(idx)
            suspect_indexes.add(idx)
            continue
        if control_ratio > 0.05:
            page_no = _table_page(table)
            if page_no is None:
                continue
            suspect_by_page[page_no].append(idx)
            suspect_indexes.add(idx)
    return suspect_by_page, suspect_indexes


def _normalize_title_candidate(raw: Optional[str]) -> Optional[str]:
    if not isinstance(raw, str):
        return None
    cleaned = " ".join(raw.strip().split())
    if not cleaned:
        return None
    printable = sum(1 for ch in cleaned if ch.isprintable())
    alnum = sum(1 for ch in cleaned if ch.isalnum())
    alpha = sum(1 for ch in cleaned if ch.isalpha())
    control = sum(1 for ch in cleaned if ch in _CONTROL_CHARS)
    if control:
        return None
    if printable / len(cleaned) < TITLE_PRINTABLE_THRESHOLD:
        return None
    if alnum / len(cleaned) < TITLE_ALNUM_THRESHOLD:
        return None
    if alpha / len(cleaned) < TITLE_ALPHA_THRESHOLD:
        return None
    if alpha < TITLE_MIN_ALPHA:
        return None
    return cleaned


def _table_bbox(table: TableItem) -> Optional[BoundingBox]:
    prov = getattr(table, "prov", None) or []
    if prov:
        entry = prov[0]
        return getattr(entry, "bbox", None)
    cluster = getattr(table, "cluster", None)
    if cluster is not None:
        return getattr(cluster, "bbox", None)
    return None


def _table_page(table: TableItem) -> Optional[int]:
    prov = getattr(table, "prov", None) or []
    if prov:
        entry = prov[0]
        return getattr(entry, "page_no", None)
    return None


def _normalize_bounds(bbox: BoundingBox) -> Tuple[float, float, float, float]:
    left = min(bbox.l, bbox.r)
    right = max(bbox.l, bbox.r)
    low = min(bbox.t, bbox.b)
    high = max(bbox.t, bbox.b)
    return left, right, low, high


def _bounds_from_bbox(bbox: Any) -> Optional[Tuple[float, float, float, float]]:
    if bbox is None:
        return None
    if hasattr(bbox, "l"):
        return _normalize_bounds(bbox)
    if isinstance(bbox, dict):
        l = bbox.get("l")
        r = bbox.get("r")
        t = bbox.get("t")
        b = bbox.get("b")
        if any(value is None for value in (l, r, t, b)):
            return None
        left = min(l, r)
        right = max(l, r)
        low = min(t, b)
        high = max(t, b)
        return left, right, low, high
    return None


def _label_name(label: Any) -> str:
    if label is None:
        return ""
    if isinstance(label, str):
        return label.lower()
    value = getattr(label, "value", None)
    if isinstance(value, str):
        return value.lower()
    try:
        return str(label).lower()
    except Exception:
        return ""


def _vertical_overlap(a: Tuple[float, float, float, float], b: Tuple[float, float, float, float], min_overlap: float = 10.0) -> bool:
    low = max(a[2], b[2])
    high = min(a[3], b[3])
    return high - low > min_overlap


def _horizontal_overlap(a: Tuple[float, float, float, float], b: Tuple[float, float, float, float], min_overlap: float = 10.0) -> bool:
    low = max(a[0], b[0])
    high = min(a[1], b[1])
    return (high - low) >= min_overlap


def _clip_cells_left_of(table: TableItem, min_x: float) -> int:
    data = getattr(table, "data", None)
    if data is None:
        return 0
    cells: List[Any] = getattr(data, "table_cells", []) or []
    if not cells:
        return 0

    kept: List[Any] = []
    removed = 0
    for cell in cells:
        bbox = getattr(cell, "bbox", None)
        if bbox is None:
            kept.append(cell)
            continue
        left = min(bbox.l, bbox.r)
        right = max(bbox.l, bbox.r)
        center_x = (left + right) / 2
        if center_x < min_x:
            removed += 1
            continue
        kept.append(cell)

    if removed:
        data.table_cells = kept
        _reindex_table_columns(data)
    return removed


def _reindex_table_columns(table_data: Any) -> None:
    cells: List[Any] = getattr(table_data, "table_cells", []) or []
    if not cells:
        table_data.num_cols = 0
        return

    unique_starts = sorted({cell.start_col_offset_idx for cell in cells})
    index_map = {old: idx for idx, old in enumerate(unique_starts)}
    max_end = 0

    for cell in cells:
        width = max(1, cell.end_col_offset_idx - cell.start_col_offset_idx)
        new_start = index_map[cell.start_col_offset_idx]
        cell.start_col_offset_idx = new_start
        cell.end_col_offset_idx = new_start + width
        max_end = max(max_end, cell.end_col_offset_idx)

    table_data.num_cols = max_end


def _table_text_sample(table: TableItem, limit: int = 4000) -> str:
    data = getattr(table, "data", None)
    if data is None:
        return ""
    grid = getattr(data, "grid", None)
    if not grid:
        return ""
    collected: List[str] = []
    total = 0
    for row in grid:
        for cell in row:
            text = getattr(cell, "text", "") or ""
            if text:
                collected.append(text)
                total += len(text)
                if total >= limit:
                    break
        if total >= limit:
            break
    return " ".join(collected)


def write_table_metadata(
    document: Any,
    target: Path,
    *,
    clipping_enabled: bool,
    clipped_cells_total: int,
    clipped_by_table: Optional[Dict[int, int]] = None,
    table_mode: str = "",
    table_cell_matching: bool = True,
    formats: Sequence[str] | None = None,
    page_metadata: Optional[Dict[int, dict[str, float]]] = None,
    garbled_tables: Optional[Set[int]] = None,
    repaired_tables: Optional[Set[int]] = None,
) -> Path:
    """Persist a detailed summary of every table in the document."""

    tables: Sequence[TableItem] = getattr(document, "tables", None) or []
    pages_meta = page_metadata or _page_metadata(document)
    doc_meta = {
        "name": getattr(document, "name", None),
        "version": getattr(document, "version", None),
        "page_count": len(pages_meta),
    }

    clipped_by_table = clipped_by_table or {}
    table_entries: List[dict[str, Any]] = []
    tables_with_clipping = 0

    for idx, table in enumerate(tables):
        data = getattr(table, "data", None)
        bbox = _table_bbox(table)
        page_no = _table_page(table)
        page_stats = pages_meta.get(page_no)
        table_stats = _table_statistics(table, data, bbox, page_stats)
        clipped = clipped_by_table.get(id(table), 0)
        if clipped:
            tables_with_clipping += 1

        table_title = _table_title(table, idx + 1, document)
        table_entries.append(
            {
                "index": idx,
                "title": table_title,
                "label": getattr(table, "label", None),
                "page": page_no,
                "self_ref": getattr(table, "self_ref", None),
                "bbox": _bbox_to_dict(bbox),
                "metrics": table_stats,
                "clipped_cells": clipped,
                "flags": {
                    "garbled_input": idx in (garbled_tables or set()),
                    "repaired_with_ocr": idx in (repaired_tables or set()),
                },
            }
        )

    summary = {
        "document": doc_meta,
        "table_processing": {
            "table_mode": table_mode,
            "table_cell_matching": table_cell_matching,
            "clip_table_overlap": clipping_enabled,
            "clipped_cells_total": clipped_cells_total,
            "tables_with_clipping": tables_with_clipping,
            "exported_formats": sorted(set(formats or [])),
            "tables_flagged_garbled": len(garbled_tables or set()),
            "tables_repaired_with_ocr": len(repaired_tables or set()),
        },
        "tables": table_entries,
    }

    payload = json.dumps(summary, ensure_ascii=False, indent=2)
    return _write_text_with_fallback(target, payload, encoding="utf-8")


def _bbox_to_dict(bbox: Optional[BoundingBox]) -> Optional[dict[str, float]]:
    if bbox is None:
        return None
    return {
        "l": bbox.l,
        "r": bbox.r,
        "t": bbox.t,
        "b": bbox.b,
        "coord_origin": str(getattr(bbox, "coord_origin", "")),
    }


def _clamp_bbox(bbox: BoundingBox, page_w: float, page_h: float) -> None:
    def clamp(value: float, minimum: float, maximum: float) -> float:
        return max(minimum, min(maximum, value))

    bbox.l = clamp(bbox.l, 0.0, page_w)
    bbox.r = clamp(bbox.r, 0.0, page_w)
    if bbox.coord_origin == CoordOrigin.TOPLEFT:
        bbox.t = clamp(bbox.t, 0.0, page_h)
        bbox.b = clamp(bbox.b, 0.0, page_h)
        if bbox.b < bbox.t:
            bbox.t, bbox.b = min(bbox.t, bbox.b), max(bbox.t, bbox.b)
    else:
        bbox.b = clamp(bbox.b, 0.0, page_h)
        bbox.t = clamp(bbox.t, 0.0, page_h)
        if bbox.t < bbox.b:
            bbox.b, bbox.t = min(bbox.b, bbox.t), max(bbox.b, bbox.t)


def _page_metadata(document: Any) -> Dict[int, dict[str, float]]:
    pages_meta: Dict[int, dict[str, float]] = {}
    pages = getattr(document, "pages", {}) or {}
    for key, page in pages.items():
        try:
            page_no = int(key)
        except Exception:
            page_no = key
        size = getattr(page, "size", None)
        if size is None:
            continue
        width = getattr(size, "width", None)
        height = getattr(size, "height", None)
        if width is None or height is None:
            continue
        pages_meta[page_no] = {
            "width": width,
            "height": height,
            "area": width * height if width and height else None,
        }
    return pages_meta


def _table_statistics(
    table: TableItem,
    data: Any,
    bbox: Optional[BoundingBox],
    page_stats: Optional[dict[str, float]],
) -> dict[str, Any]:
    num_rows = getattr(data, "num_rows", None)
    num_cols = getattr(data, "num_cols", None)
    grid = getattr(data, "grid", None) or []
    cells = getattr(data, "table_cells", None) or []
    total_cells = len(cells) if cells else None
    header_rows = _count_header_rows(grid)
    header_cols = _count_header_columns(grid, num_cols)
    bbox_metrics = _bbox_metrics(bbox, page_stats)
    merged_cells = any(
        getattr(cell, "row_span", 1) > 1 or getattr(cell, "col_span", 1) > 1
        for cell in cells
    )
    density = None
    if num_rows and num_cols and cells:
        density = len(cells) / float(num_rows * num_cols)

    return {
        "num_rows": num_rows,
        "num_cols": num_cols,
        "header_rows": header_rows,
        "header_columns": header_cols,
        "total_cells": total_cells,
        "cells_with_bbox": _count_cells_with_bbox(cells),
        "data_density": density,
        "has_merged_cells": merged_cells,
        "avg_column_widths": _average_column_widths(grid, num_cols),
        "avg_row_heights": _average_row_heights(grid, num_rows),
        "bbox_metrics": bbox_metrics,
    }


def _count_header_rows(grid: Sequence[Sequence[Any]]) -> int:
    count = 0
    for row in grid:
        if any(getattr(cell, "column_header", False) for cell in row):
            count += 1
        else:
            break
    return count


def _count_header_columns(grid: Sequence[Sequence[Any]], num_cols: Optional[int]) -> Optional[int]:
    if not num_cols:
        return None
    header_cols = set()
    for row in grid:
        for idx, cell in enumerate(row):
            if getattr(cell, "column_header", False):
                header_cols.add(idx)
    return len(header_cols) if header_cols else 0


def _count_cells_with_bbox(cells: Sequence[Any]) -> Optional[int]:
    if not cells:
        return None
    return sum(1 for cell in cells if getattr(cell, "bbox", None) is not None)


def _average_column_widths(grid: Sequence[Sequence[Any]], num_cols: Optional[int]) -> Optional[List[Optional[float]]]:
    if not num_cols or not grid:
        return None
    totals = [0.0] * num_cols
    counts = [0] * num_cols
    for row in grid:
        for idx in range(min(len(row), num_cols)):
            bbox = getattr(row[idx], "bbox", None)
            if bbox is None:
                continue
            width = abs(bbox.r - bbox.l)
            totals[idx] += width
            counts[idx] += 1
    averages = []
    for total, count in zip(totals, counts):
        averages.append(total / count if count else None)
    return averages


def _average_row_heights(grid: Sequence[Sequence[Any]], num_rows: Optional[int]) -> Optional[List[Optional[float]]]:
    if not num_rows or not grid:
        return None
    heights = []
    for row in grid:
        row_heights = [
            abs(cell.bbox.t - cell.bbox.b)
            for cell in row
            if getattr(cell, "bbox", None) is not None
        ]
        heights.append(sum(row_heights) / len(row_heights) if row_heights else None)
    return heights if any(h is not None for h in heights) else None


def _bbox_metrics(bbox: Optional[BoundingBox], page_stats: Optional[dict[str, float]]) -> Optional[dict[str, float]]:
    if bbox is None:
        return None
    width = abs(bbox.r - bbox.l)
    height = abs(bbox.t - bbox.b)
    area = width * height
    coverage = None
    if page_stats and page_stats.get("area"):
        coverage = area / page_stats["area"] if page_stats["area"] else None
    return {
        "width": width,
        "height": height,
        "area": area,
        "coverage": coverage,
    }


def _table_bbox_tuple(table: TableItem) -> Optional[Tuple[float, float, float, float]]:
    prov = getattr(table, "prov", None) or []
    if not prov:
        return None
    bbox = getattr(prov[0], "bbox", None)
    if bbox is None:
        return None
    bounds = _normalize_bounds(bbox)
    return bounds[0], bounds[2], bounds[1], bounds[3]


def _iou(a: Tuple[float, float, float, float], b: Tuple[float, float, float, float]) -> float:
    ax0, ay0, ax1, ay1 = a
    bx0, by0, bx1, by1 = b
    inter_x0 = max(ax0, bx0)
    inter_y0 = max(ay0, by0)
    inter_x1 = min(ax1, bx1)
    inter_y1 = min(ay1, by1)
    inter_w = inter_x1 - inter_x0
    inter_h = inter_y1 - inter_y0
    if inter_w <= 0 or inter_h <= 0:
        return 0.0
    inter_area = inter_w * inter_h
    area_a = max((ax1 - ax0) * (ay1 - ay0), 1e-6)
    area_b = max((bx1 - bx0) * (by1 - by0), 1e-6)
    union = area_a + area_b - inter_area
    return inter_area / max(union, 1e-6)


def _copy_table_contents(target: TableItem, source: TableItem) -> None:
    clone = getattr(source, "model_copy", None)
    if callable(clone):
        replacement = source.model_copy(deep=True)
        target.data = replacement.data
        target.captions = replacement.captions
        target.references = replacement.references
        target.footnotes = replacement.footnotes
        target.annotations = replacement.annotations
        target.meta = replacement.meta
        target.prov = replacement.prov
        target.label = replacement.label
    else:  # pragma: no cover - fallback for stub objects
        target.data = getattr(source, "data", None)
        target.captions = getattr(source, "captions", [])
        target.references = getattr(source, "references", [])
        target.footnotes = getattr(source, "footnotes", [])
        target.annotations = getattr(source, "annotations", [])
        target.meta = getattr(source, "meta", None)
        target.prov = getattr(source, "prov", [])
        target.label = getattr(source, "label", None)


def merge_repaired_tables(
    base_document: Any,
    repair_document: Any,
    table_indexes: List[int],
    page_no: int,
) -> Set[int]:
    """Replace the specified tables on a page with the repair document's tables."""

    replaced: Set[int] = set()
    repair_tables = [
        table for table in getattr(repair_document, "tables", []) or [] if _table_page(table) == page_no
    ]
    used: Set[int] = set()
    for table_idx in table_indexes:
        try:
            target_table = base_document.tables[table_idx]
        except IndexError:
            continue
        target_bbox = _table_bbox_tuple(target_table)
        if target_bbox is None:
            continue
        best_score = 0.0
        best_j = None
        for j, candidate in enumerate(repair_tables):
            if j in used:
                continue
            candidate_bbox = _table_bbox_tuple(candidate)
            if candidate_bbox is None:
                continue
            score = _iou(target_bbox, candidate_bbox)
            if score > best_score:
                best_score = score
                best_j = j
        if best_j is not None and best_score > 0.1:
            _copy_table_contents(target_table, repair_tables[best_j])
            used.add(best_j)
            replaced.add(table_idx)
    return replaced


def export_tables_to_xlsx(document: Any, target: Path) -> Optional[Path]:
    tables: Sequence[TableItem] = getattr(document, "tables", None) or []
    if not tables:
        return None

    try:
        from openpyxl import Workbook
    except ImportError:  # pragma: no cover
        raise RuntimeError("openpyxl is required for Excel export but is not installed.")

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    tables_by_page: Dict[Optional[int], int] = defaultdict(int)
    for idx, table in enumerate(tables, start=1):
        data = getattr(table, "data", None)
        grid = getattr(data, "grid", None)
        page_no = _table_page(table)
        tables_by_page[page_no] += 1
        local_idx = tables_by_page[page_no]
        title = _table_title(table, local_idx, document)
        sheet = wb.create_sheet(_sheet_name(page_no, local_idx, title))
        if not grid:
            continue
        for row_idx, row in enumerate(grid, start=1):
            for col_idx, cell in enumerate(row, start=1):
                text = ""
                if cell is not None:
                    text = getattr(cell, "text", "")
                sheet.cell(row=row_idx, column=col_idx, value=_sanitize_excel_text(text))

    return _save_workbook_with_fallback(wb, target)


def _sheet_name(page_no: Optional[int], local_index: int, title: Optional[str]) -> str:
    page_token = f"P{page_no}" if page_no is not None else "PXX"
    prefix = f"{page_token}_T{local_index:02d}"
    cleaned = _sanitize_sheet_title(title or "")
    max_len = 31
    if cleaned:
        available = max_len - len(prefix) - 1
        if available > 0:
            cleaned = cleaned[:available]
            return f"{prefix}_{cleaned}"
    name = prefix[:max_len]
    return name or f"Table{local_index:02d}"


def _sanitize_sheet_title(title: str) -> str:
    title = title.replace("\n", " ").replace("\r", " ")
    title = re.sub(r"[\[\]:*?/\\]", "", title)
    return title.strip()


def _sanitize_excel_text(text: str) -> str:
    if not text:
        return ""
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    text = text.replace("\r", " ").replace("\n", " ")
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    if text and text[0] in ("=", "+", "-", "@"):
        text = "\u200B" + text
    return text


def _table_title(table: TableItem, index: int, document: Any) -> Optional[str]:
    caption_text = None
    try:
        caption_text = table.caption_text(document)
    except Exception:
        caption_text = None
    candidate = _normalize_title_candidate(caption_text)
    if candidate:
        return candidate

    label = getattr(table, "label", None)
    if isinstance(label, str):
        label_text = _normalize_title_candidate(label)
        if label_text and _label_name(label) not in {"table"}:
            return label_text

    heading = _heading_title_from_document(table, document)
    if heading:
        return heading

    data = getattr(table, "data", None)
    merged_header = _merged_header_title(data)
    if merged_header:
        return merged_header

    grid = getattr(data, "grid", None)
    if grid and grid[0] and getattr(grid[0][0], "text", ""):
        first_text = _normalize_title_candidate(grid[0][0].text)
        if first_text and first_text.lower().startswith("table"):
            return first_text

    return f"Table {index:02d}"


def _table_title_from_json(table: Dict[str, Any], index: int, document: Optional[Dict[str, Any]] = None) -> Optional[str]:
    captions = table.get("captions") or []
    if captions:
        caption_text = captions[0]
        if isinstance(caption_text, dict):
            caption_text = caption_text.get("text")
        candidate = _normalize_title_candidate(caption_text)
        if candidate:
            return candidate

    label = table.get("label")
    if isinstance(label, str):
        label_text = _normalize_title_candidate(label)
        if label_text and label_text.lower() != "table":
            return label_text

    if document:
        heading = _heading_title_from_json(table, document)
        if heading:
            return heading

    data = table.get("data") or {}
    merged_header = _merged_header_title_from_json(data)
    if merged_header:
        return merged_header

    grid = data.get("grid") or []
    if grid and grid[0] and isinstance(grid[0][0], dict):
        text = _normalize_title_candidate(grid[0][0].get("text"))
        if text and text.lower().startswith("table"):
            return text

    page_no = None
    prov = table.get("prov") or []
    if prov:
        entry = prov[0]
        page_no = entry.get("page_no")
    if page_no is not None:
        return f"Page {page_no} Table {index:02d}"
    return f"Table {index:02d}"


def _heading_title_from_document(table: TableItem, document: Any) -> Optional[str]:
    headings = getattr(document, "texts", None) or []
    if not headings:
        return None
    bbox = _table_bbox(table)
    if bbox is None:
        return None
    table_bounds = _normalize_bounds(bbox)
    page_no = _table_page(table)
    best_text: Optional[str] = None
    best_gap = float("inf")
    for heading in headings:
        label = getattr(heading, "label", None)
        label_name = _label_name(label)
        if label_name not in HEADING_LABELS:
            continue
        prov = getattr(heading, "prov", None) or []
        if not prov:
            continue
        entry = prov[0]
        if page_no is not None and getattr(entry, "page_no", None) != page_no:
            continue
        heading_bbox = getattr(entry, "bbox", None)
        if heading_bbox is None:
            continue
        heading_bounds = _normalize_bounds(heading_bbox)
        if not _horizontal_overlap(heading_bounds, table_bounds, min_overlap=20.0):
            continue
        gap = heading_bounds[2] - table_bounds[3]
        if gap < -20.0 or gap > 150.0:
            continue
        text_value = getattr(heading, "text", None) or getattr(heading, "orig", None)
        stripped = _normalize_title_candidate(text_value)
        if not stripped:
            continue
        if gap < best_gap:
            best_gap = gap
            best_text = stripped
    return best_text


def _heading_title_from_json(table: Dict[str, Any], document: Dict[str, Any]) -> Optional[str]:
    headings: Sequence[Dict[str, Any]] = document.get("texts") or []
    if not headings:
        return None
    prov = table.get("prov") or []
    bbox = prov[0].get("bbox") if prov else None
    table_bounds = _bounds_from_bbox(bbox)
    if table_bounds is None:
        return None
    page_no = prov[0].get("page_no") if prov else None
    best_text: Optional[str] = None
    best_gap = float("inf")
    for heading in headings:
        label = (heading.get("label") or "").lower()
        if label not in HEADING_LABELS:
            continue
        heading_prov = heading.get("prov") or []
        if not heading_prov:
            continue
        if page_no is not None and heading_prov[0].get("page_no") != page_no:
            continue
        heading_bounds = _bounds_from_bbox(heading_prov[0].get("bbox"))
        if heading_bounds is None:
            continue
        if not _horizontal_overlap(heading_bounds, table_bounds, min_overlap=20.0):
            continue
        gap = heading_bounds[2] - table_bounds[3]
        if gap < -20.0 or gap > 150.0:
            continue
        text_value = heading.get("text") or heading.get("orig")
        stripped = _normalize_title_candidate(text_value)
        if not stripped:
            continue
        if gap < best_gap:
            best_gap = gap
            best_text = stripped
    return best_text


def _cell_col_span(cell: Any) -> Optional[int]:
    if cell is None:
        return None
    span = getattr(cell, "col_span", None)
    if isinstance(span, int) and span > 0:
        return span
    start = getattr(cell, "start_col_offset_idx", None)
    end = getattr(cell, "end_col_offset_idx", None)
    if isinstance(start, int) and isinstance(end, int) and end > start:
        return end - start
    if isinstance(cell, dict):
        dict_span = cell.get("col_span")
        if isinstance(dict_span, int) and dict_span > 0:
            return dict_span
        start = cell.get("start_col_offset_idx")
        end = cell.get("end_col_offset_idx")
        if isinstance(start, int) and isinstance(end, int) and end > start:
            return end - start
    return None


def _merged_header_title(data: Any) -> Optional[str]:
    if data is None:
        return None
    grid = getattr(data, "grid", None)
    if not grid:
        return None
    first_row = grid[0]
    if not first_row:
        return None
    first_cell = first_row[0]
    num_cols = getattr(data, "num_cols", None)
    span = _cell_col_span(first_cell)
    if isinstance(num_cols, int) and isinstance(span, int) and num_cols > 0 and span >= num_cols:
        text = getattr(first_cell, "text", "") or ""
    else:
        texts = {getattr(cell, "text", "").strip() for cell in first_row if getattr(cell, "text", "").strip()}
        text = texts.pop() if len(texts) == 1 else ""
    return _normalize_title_candidate(text)


def _merged_header_title_from_json(data: Dict[str, Any]) -> Optional[str]:
    if not data:
        return None
    grid = data.get("grid") or []
    if not grid:
        return None
    first_row = grid[0]
    if not first_row:
        return None
    first_cell = first_row[0]
    num_cols = data.get("num_cols")
    span = _cell_col_span(first_cell)
    if isinstance(num_cols, int) and isinstance(span, int) and num_cols > 0 and span >= num_cols:
        text = first_cell.get("text", "")
    else:
        texts = {str(cell.get("text") or "").strip() for cell in first_row if isinstance(cell, dict) and (cell.get("text") or "").strip()}
        text = texts.pop() if len(texts) == 1 else ""
    return _normalize_title_candidate(text)


def _next_available_path(path: Path) -> Path:
    base = Path(path)
    for idx in itertools.count(1):
        candidate = base.with_name(f"{base.stem} ({idx}){base.suffix}")
        if not candidate.exists():
            return candidate


def _write_text_with_fallback(target: Path, content: str, encoding: str = "utf-8") -> Path:
    attempt = Path(target)
    tried = 0
    while True:
        try:
            attempt.parent.mkdir(parents=True, exist_ok=True)
            attempt.write_text(content, encoding=encoding)
            if tried:
                logger.warning("Base file '%s' was locked; wrote to '%s' instead.", target.name, attempt.name)
            return attempt
        except PermissionError:
            tried += 1
            attempt = _next_available_path(target)


def _save_workbook_with_fallback(workbook, target: Path) -> Path:
    attempt = Path(target)
    tried = 0
    while True:
        try:
            attempt.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(attempt)
            if tried:
                logger.warning("Base workbook '%s' was locked; saved to '%s' instead.", target.name, attempt.name)
            return attempt
        except PermissionError:
            tried += 1
            attempt = _next_available_path(target)


__all__ = [
    "ensure_tableformer_structure",
    "clip_overlapping_table_cells",
    "write_table_metadata",
    "export_tables_to_xlsx",
    "sanitize_table_bboxes",
    "detect_garbled_tables",
    "merge_repaired_tables",
    "_write_text_with_fallback",
    "_save_workbook_with_fallback",
    "_table_title_from_json",
]
